"""Feature 5 — deterministic recovery projection → XLSX (Excel) exporter.

Banks exchange the *numbers* of a 経営改善計画 in Excel, so the P&L recovery
curve (the month-by-month uplift / 経常利益 / EWS grid) is the natural XLSX
export — not the prose document (that is PDF/DOCX).

Numeric-preservation invariant
------------------------------
Every yen figure and EWS value written to the workbook is taken **directly**
from the :class:`~app.backend.analysis.pnl_recovery.RecoveryProjection` integer
/ float fields. Nothing is re-derived or reformatted: the cells carry the exact
values the deterministic projection produced. ``openpyxl`` is the only
dependency.
"""

from __future__ import annotations

import io

from openpyxl import Workbook

from app.backend.analysis.pnl_recovery import RecoveryProjection
from app.backend.export._filenames import safe_filename_stem
from app.shared.constants import EWS_SUBSTANDARD

__all__ = ["build_recovery_xlsx", "xlsx_filename"]

#: Yen number format (no decimals, thousands separators, negatives in red).
_YEN_FMT = "#,##0;[Red]-#,##0"
#: EWS number format (two decimals).
_EWS_FMT = "0.00"


def build_recovery_xlsx(recovery: RecoveryProjection) -> bytes:
    """Build an ``.xlsx`` of the recovery projection (the P&L bridge grid).

    Writes a header block (annual uplift, full monthly uplift, ramp, recovery
    verdict) and a per-month table (month, booked uplift, projected 経常利益,
    EWS). Every figure is copied verbatim from the projection's integer/float
    fields — no value is re-derived or reformatted.

    Args:
        recovery: The deterministic recovery projection.

    Returns:
        The generated ``.xlsx`` file as bytes.
    """
    workbook = Workbook()
    # Workbook.active is Optional in openpyxl's typing; a fresh workbook always
    # has one, but guard explicitly so this can never be a NoneType crash and
    # passes strict type-checking.
    sheet = workbook.active
    if sheet is None:
        sheet = workbook.create_sheet()
    sheet.title = "損益計画"

    # --- Header block (labels + values) ---
    sheet["A1"] = "損益計画（Recovery projection）"
    sheet["A2"] = "期待経常利益改善 / 年 (Annual uplift)"
    sheet["B2"] = int(recovery.annual_uplift)
    sheet["B2"].number_format = _YEN_FMT
    sheet["A3"] = "月次換算 (Full monthly uplift)"
    sheet["B3"] = int(recovery.full_monthly_uplift)
    sheet["B3"].number_format = _YEN_FMT
    sheet["A4"] = "段階導入月数 (Ramp months)"
    sheet["B4"] = int(recovery.ramp_months)
    sheet["A5"] = "正常化見込 (Projected normalisation)"
    if recovery.recovery_month_index is not None:
        sheet["B5"] = f"{recovery.recovery_month_index}ヶ月目 (EWS < {int(EWS_SUBSTANDARD)})"
    else:
        sheet["B5"] = f"{len(recovery.months)}ヶ月以内には未達 (EWS ≥ {int(EWS_SUBSTANDARD)})"

    # --- Table header ---
    header_row = 7
    headers = [
        "月 (Month)",
        "月次改善額 (Uplift)",
        "経常利益 (Keijo Rieki)",
        "EWS",
        "正常化 (Recovered)",
    ]
    for col, text in enumerate(headers, start=1):
        sheet.cell(row=header_row, column=col, value=text)

    # --- Per-month rows (values copied verbatim from the projection) ---
    for offset, month in enumerate(recovery.months, start=1):
        row = header_row + offset
        sheet.cell(row=row, column=1, value=int(month.month_index))
        cell_uplift = sheet.cell(row=row, column=2, value=int(month.monthly_uplift))
        cell_uplift.number_format = _YEN_FMT
        cell_keijo = sheet.cell(row=row, column=3, value=int(month.keijo_rieki))
        cell_keijo.number_format = _YEN_FMT
        cell_ews = sheet.cell(row=row, column=4, value=float(month.ews_score))
        cell_ews.number_format = _EWS_FMT
        sheet.cell(row=row, column=5, value="✓" if month.recovered else "")

    # Reasonable column widths for readability (cosmetic only).
    for col_letter, width in {"A": 26, "B": 22, "C": 24, "D": 10, "E": 14}.items():
        sheet.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def xlsx_filename(company_or_code: str) -> str:
    """Return a safe ``recovery_<name>.xlsx`` filename.

    Sanitises the name for cross-OS downloads (collapses path separators /
    whitespace and strips Windows-illegal characters ``: * ? " < > |`` and
    control chars); falls back to ``recovery`` when the name sanitises to empty.
    """
    stem = safe_filename_stem(company_or_code or "", fallback="recovery")
    return f"recovery_{stem}.xlsx"

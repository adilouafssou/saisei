"""Tests for the Feature 5 recovery → XLSX exporter and the Keikakusho section.

The load-bearing invariant is numeric preservation: every yen / EWS value in
the workbook (and in the rendered Markdown section) must equal the exact value
the deterministic projection produced — nothing re-derived or reformatted.
"""

from __future__ import annotations

import datetime as dt
import io

from app.backend.analysis.pnl_recovery import project_recovery
from app.backend.export.recovery_xlsx import build_recovery_xlsx, xlsx_filename
from app.backend.nodes.kaizen_generation import render_keikakusho
from app.backend.state import Strategy
from app.shared.models.accounting import TrialBalance
from openpyxl import load_workbook


def _declining_history() -> list[TrialBalance]:
    rows: list[TrialBalance] = []
    for i in range(12):
        sales = 150_000_000 - i * 2_500_000
        cogs = int(sales * (0.80 + i * 0.005))
        rows.append(
            TrialBalance(
                period=dt.date(2025, 1, 1) + dt.timedelta(days=30 * i),
                uriage=sales,
                uriage_genka=cogs,
                hanbaihi=20_000_000,
            )
        )
    return rows


def test_xlsx_is_valid_zip_and_has_sheet() -> None:
    proj = project_recovery(_declining_history(), 120_000_000, horizon_months=36)
    data = build_recovery_xlsx(proj)
    assert data[:2] == b"PK"
    wb = load_workbook(io.BytesIO(data))
    assert "損益計画" in wb.sheetnames


def test_xlsx_cells_match_projection_values_verbatim() -> None:
    proj = project_recovery(
        _declining_history(), 120_000_000, horizon_months=36, stop_at_recovery=False
    )
    wb = load_workbook(io.BytesIO(build_recovery_xlsx(proj)))
    sheet = wb.active
    # Header values copied verbatim.
    assert sheet["B2"].value == proj.annual_uplift
    assert sheet["B3"].value == proj.full_monthly_uplift
    assert sheet["B4"].value == proj.ramp_months
    # First data row (table header at row 7, first month at row 8).
    first = proj.months[0]
    assert sheet.cell(row=8, column=1).value == first.month_index
    assert sheet.cell(row=8, column=2).value == first.monthly_uplift
    assert sheet.cell(row=8, column=3).value == first.keijo_rieki
    assert abs(sheet.cell(row=8, column=4).value - first.ews_score) < 1e-9


def test_xlsx_filename_is_safe() -> None:
    assert xlsx_filename("テスト 製造") == "recovery_テスト_製造.xlsx"
    assert xlsx_filename("") == "recovery_recovery.xlsx"


def test_render_keikakusho_without_recovery_is_unchanged() -> None:
    """When no recovery is passed, the draft must NOT contain a section 4."""
    latest = _declining_history()[-1]
    strategy = Strategy(title="価格転嫁", rationale="...", expected_keijo_uplift=54_000_000)
    draft = render_keikakusho(
        company_name="テスト社",
        hojin_bango="1234567890123",
        fsa_kanji="要注意先",
        latest=latest,
        strategy=strategy,
        working_capital_gap=-5_000_000,
    )
    assert "## 4." not in draft
    assert "損益計画" not in draft


def test_render_keikakusho_with_recovery_appends_section_4() -> None:
    history = _declining_history()
    strategy = Strategy(title="価格転嫁", rationale="...", expected_keijo_uplift=120_000_000)
    proj = project_recovery(history, int(strategy.expected_keijo_uplift))
    draft = render_keikakusho(
        company_name="テスト社",
        hojin_bango="1234567890123",
        fsa_kanji="要注意先",
        latest=history[-1],
        strategy=strategy,
        working_capital_gap=-5_000_000,
        recovery=proj,
    )
    assert "## 4. 損益計画（Recovery projection）" in draft
    # The annual uplift figure appears verbatim in the section.
    assert "120,000,000" in draft
